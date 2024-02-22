import os

from aiogram import types, F, Router, Bot
from typing import Optional
from aiogram.filters.callback_data import CallbackData
from aiogram.types import FSInputFile
from openpyxl import Workbook

from bot.dop_func import clean_tmp_files
from bot.keyboards.staff import main_klasses_kb, KlassCallbackFactory, klass_cancel_kb, \
    klass_choice_kb, current_klass_kb, klass_delete_confirm, klass_parallel_choice_kb
from aiogram.fsm.context import FSMContext

from bot.states.staff_states import Content_add
from db import Parallel, Klass
from settings import tmp_files

klasses_router = Router()


@klasses_router.callback_query(F.data == 'klasses')
async def klasses_main(clb: types.CallbackQuery, state: FSMContext, bot: Bot) -> None:
    await state.clear()
    reply_text = 'Выберите действие:'
    await bot.edit_message_text(chat_id=clb.message.chat.id, message_id=clb.message.message_id, text=reply_text,
                                reply_markup=main_klasses_kb())


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "add_klass"))
async def callbacks_klass_add_fab(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                  state: FSMContext):
    parallels = Parallel.select()
    text = 'Выберите параллель:'
    await callback.message.edit_text(text=text,
                                     reply_markup=klass_parallel_choice_kb(parallels, action='choice_for_new_klass'))


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "choice_for_new_klass"))
async def callbacks_klass_add2_fab(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                   state: FSMContext):
    parallel_id = callback_data.id
    await state.set_state(Content_add.klass_add)
    await state.update_data(msg_id=callback.message.message_id)
    await state.update_data(parallel_id=parallel_id)
    text = 'Введите название класса'
    await callback.message.edit_text(text=text, reply_markup=klass_cancel_kb())


@klasses_router.message(Content_add.klass_add)
async def klass_add(msg: types.Message, state: FSMContext, bot: Bot):
    klass_name = msg.text.title()
    contex_data = await state.get_data()
    msg_id = contex_data.get('msg_id')
    parallel_id = contex_data.get('parallel_id')
    parallel = Parallel.get(id=parallel_id)
    klass = Klass.create(name=klass_name, parallel=parallel)
    await msg.delete()
    await state.clear()
    text = f'Создан класс № {klass.id}) {klass.name}'
    await bot.edit_message_text(chat_id=msg.chat.id, message_id=msg_id, text=text, reply_markup=main_klasses_kb())


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "search_klass"))
async def callbacks_klass_search_fab(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                     state: FSMContext):
    parallels = Parallel.select()
    text = 'Выберите параллель:'
    await callback.message.edit_text(text=text,
                                     reply_markup=klass_parallel_choice_kb(parallels, action='choice_for_search'))


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "choice_for_search"))
async def callbacks_klass_search2_fab(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                      state: FSMContext):
    parallel_id = callback_data.id
    parallel = Parallel.get(parallel_id)
    klasses = parallel.klasses
    text = 'Выберите класс:'
    await callback.message.edit_text(text=text, reply_markup=klass_choice_kb(klasses))


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "open"))
async def callbacks_author_open(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                state: FSMContext):
    klass_id = callback_data.id
    klass = Klass.get(id=klass_id)
    text = f'Класс: {klass.name} \n[{klass.parallel.name} параллель]'
    await callback.message.edit_text(text=text, reply_markup=current_klass_kb(klass_id))


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "download_codes"))
async def callbacks_download_codes_fab(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                       state: FSMContext, bot: Bot):
    def create_invite_codes(file_path, klass):
        wb = Workbook()
        ws = wb.active

        for i, student in enumerate(klass.students, start=1):
            ws.cell(row=i, column=1).value = student.fio
            ws.cell(row=i, column=2).value = student.invite_code

        wb.save(file_path)
        wb.close()

    klass_id = callback_data.id
    klass = Klass.get(id=klass_id)

    file_name = f'Список кодов {klass.name} класс.xlsx'
    path = os.path.join(tmp_files, file_name)
    create_invite_codes(file_path=path, klass=klass)
    await callback.message.answer_document(FSInputFile(path))
    clean_tmp_files()
    await callback.message.delete()
    text = f'Класс: {klass.name} \n[{klass.parallel.name} параллель]'
    await callback.message.answer(text=text, reply_markup=current_klass_kb(klass_id))


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "edit_name"))
async def callbacks_klass_edit_name(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                    state: FSMContext):
    klass_id = callback_data.id
    await state.set_state(Content_add.klass_edit_name)
    await state.update_data(msg_id=callback.message.message_id)
    await state.update_data(klass_id=klass_id)
    text = callback.message.text + '\n🆕Введите новое название класса:'
    await callback.message.edit_text(text=text, reply_markup=klass_cancel_kb())


@klasses_router.message(Content_add.klass_edit_name)
async def klass_edit_name(msg: types.Message, state: FSMContext, bot: Bot):
    new_klass_name = msg.text.upper()
    contex_data = await state.get_data()
    msg_id = contex_data.get('msg_id')
    klass_id = contex_data.get('klass_id')
    klass = Klass.get(id=klass_id)
    klass.name = new_klass_name
    klass.save()
    await msg.delete()
    await state.clear()
    text = f'Класс: 🆕{klass.name} \n[{klass.parallel.name} параллель]'
    await bot.edit_message_text(chat_id=msg.chat.id, message_id=msg_id, text=text,
                                reply_markup=current_klass_kb(klass_id))


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "edit_parallel"))
async def callbacks_klass_edit_parallel(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                        state: FSMContext):
    klass_id = callback_data.id
    await state.set_state(Content_add.klass_edit_parallel)
    await state.update_data(klass_id=klass_id)
    text = callback.message.text + '\n🆕Выберите новую параллель для класса:'
    parallels = Parallel.select()
    await callback.message.edit_text(text=text,
                                     reply_markup=klass_parallel_choice_kb(parallels, action='choice_paral_for_edit'))


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "choice_paral_for_edit"))
async def callbacks_klass_edit_parallel(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                        state: FSMContext):
    parallel_id = callback_data.id
    parallel = Parallel.get(id=parallel_id)
    contex_data = await state.get_data()
    klass_id = contex_data.get('klass_id')
    klass = Klass.get(id=klass_id)
    klass.parallel = parallel
    klass.save()
    text = f'Класс: {klass.name} \n[🆕{klass.parallel.name} параллель]'
    await callback.message.edit_text(text=text, reply_markup=current_klass_kb(klass_id))


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "delete"))
async def callbacks_klass_delete(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                 state: FSMContext):
    klass_id = callback_data.id
    text = callback.message.text + '\n❗️Уверены, что хотите удалить класс?'
    await callback.message.edit_text(text=text, reply_markup=klass_delete_confirm(klass_id))


@klasses_router.callback_query(KlassCallbackFactory.filter(F.action == "delete_confirmed"))
async def callbacks_klass_delete_confirmed(callback: types.CallbackQuery, callback_data: KlassCallbackFactory,
                                           state: FSMContext):
    klass_id = callback_data.id
    klass = Klass.get(id=klass_id)
    klass_name = klass.name
    klass.delete_instance()
    text = f'Класс {klass_name} удалён.'
    await callback.message.edit_text(text=text, reply_markup=main_klasses_kb())
