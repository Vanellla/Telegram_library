import os
import datetime
from pprint import pprint

from openpyxl import load_workbook

from aiogram import types, F, Router, Bot
from typing import Optional
from aiogram.filters.callback_data import CallbackData

from bot.keyboards.staff import main_students_kb, StudentCallbackFactory, student_parallel_choice_kb, \
    student_cancel_kb, student_search_type_kb, student_search_result, current_student_kb, student_klass_choice_kb, \
    student_delete_confirm, student_search_result_by_fio_kb, student_book_issuance, back_to_student
from aiogram.fsm.context import FSMContext

from bot.states.staff_states import Content_add
from db import Parallel, Klass, Student, Book, Book_movement, Staff
from qr import DecodeQR
from settings import tmp_files, qr_dir

student_router = Router()


@student_router.callback_query(F.data == 'students')
async def students_main(clb: types.CallbackQuery, state: FSMContext, bot: Bot) -> None:
    await state.clear()
    reply_text = 'Выберите действие:'
    await clb.message.edit_text(text=reply_text,
                                reply_markup=main_students_kb())


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "add_student"))
async def callbacks_student_add_step1_fab(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                          state: FSMContext):
    parallels = Parallel.select()
    text = '1️⃣Выберите параллель:'
    await callback.message.edit_text(text=text,
                                     reply_markup=student_parallel_choice_kb(parallels,
                                                                             action='choice_for_add_student'))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "choice_for_add_student"))
async def callbacks_student_add_step2_fab(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                          state: FSMContext):
    parallel_id = callback_data.id
    parallel = Parallel.get(id=parallel_id)
    klasses = parallel.klasses
    text = '2️⃣Выберите класс'
    await callback.message.edit_text(text=text,
                                     reply_markup=student_klass_choice_kb(klasses, action='choice_for_new_student'))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "choice_for_new_student"))
async def callbacks_student_add_step3_fab(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                          state: FSMContext):
    klass_id = callback_data.id
    await state.set_state(Content_add.student_add)
    await state.update_data(msg_id=callback.message.message_id)
    await state.update_data(klass_id=klass_id)
    text = '3️⃣ Перечислите новых учеников одним из двух способов:\n'
    text += '1) Напишите ФИО учеников через запятую в чат,\n'
    text += '2) Пришлите в чат файл таблицы в формате xlsx, где в каждой строке ФИО нового ученика'
    await callback.message.edit_text(text=text, reply_markup=student_cancel_kb())


@student_router.message(Content_add.student_add, F.text)
async def student_add_step4(msg: types.Message, state: FSMContext, bot: Bot):
    students = msg.text.title()
    contex_data = await state.get_data()
    msg_id = contex_data.get('msg_id')
    klass_id = contex_data.get('klass_id')
    klass = Klass.get(id=klass_id)
    await msg.delete()
    students_parsed = [student.strip() for student in students.split(',')]
    validation = [student for student in students_parsed if len(student.split()) < 2 or len(student.split()) > 4]
    if len(validation) > 0:
        text = 'Вы уверены что ввели ФИО правильно?\nПроверьте следующие записи:'
        # text += 'Пример: Иванов Алексей Владимирович, Петров Николай Иванович, и т.д.'
        for student in validation:
            text += student + '\n'
        text += 'Пришлите исправленный текст'
        await bot.edit_message_text(chat_id=msg.chat.id, message_id=msg_id, text=text, reply_markup=student_cancel_kb())
    await state.clear()
    for student in students_parsed:
        Student.create(fio=student, klass=klass)
    text = f'Ученики добавлены в {klass.name} класс'
    await bot.edit_message_text(chat_id=msg.chat.id, message_id=msg_id, text=text, reply_markup=main_students_kb())


@student_router.message(Content_add.student_add, F.document)  # надо доделать
async def student_add_step3_by_xlsx(msg: types.Message, state: FSMContext, bot: Bot):
    def read_xlsx(file_path, klass):
        wb = load_workbook(file_path)
        ws = wb.active
        for row in range(1, 50):
            for col in range(1, 20):
                t = ws.cell(row=row, column=col).value
                if t != None and len(t) > 5:
                    Student.create(fio=t, klass=klass)
        wb.close()
        folder = tmp_files
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
            except:
                pass

    contex_data = await state.get_data()
    msg_id = contex_data.get('msg_id')
    klass_id = contex_data.get('klass_id')
    klass = Klass.get(id=klass_id)
    file_name = msg.document.file_name
    path = os.path.join(tmp_files, file_name)
    await bot.download(msg.document, path)
    read_xlsx(file_path=path, klass=klass)

    await msg.delete()
    await state.clear()

    text = f'Ученики добавлены в {klass.name} класс'
    await bot.edit_message_text(chat_id=msg.chat.id, message_id=msg_id, text=text, reply_markup=main_students_kb())


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "search_student"))
async def callbacks_student_search_fab(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                       state: FSMContext):
    text = 'Выберите способ поиска:'
    await callback.message.edit_text(text=text, reply_markup=student_search_type_kb())


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "search_by_fio"))
async def callbacks_student_search_fab(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                       state: FSMContext):
    # klass_id = callback_data.id
    await state.set_state(Content_add.student_search_by_fio)
    await state.update_data(msg_id=callback.message.message_id)
    text = 'Введите Фамилию ил Фамилию Имя (у имени можно первые буквы):'
    await callback.message.edit_text(text=text, reply_markup=student_cancel_kb())


@student_router.message(Content_add.student_search_by_fio)
async def student_search_by_fio(msg: types.Message, state: FSMContext, bot: Bot):
    fio = msg.text.title()
    contex_data = await state.get_data()
    msg_id = contex_data.get('msg_id')
    students = Student.select().where(Student.fio.contains(fio))
    if students.count() > 15:
        text = 'Результат поиска очень большой.\nПопробуйте указать более точный запрос.'
        await bot.edit_message_text(chat_id=msg.chat.id, message_id=msg_id, text=text, reply_markup=student_cancel_kb())
    await msg.delete()
    await state.clear()
    if students.count() > 0:
        text = 'Вот что удалось найти в базе учеников:'
        await bot.edit_message_text(chat_id=msg.chat.id, message_id=msg_id, text=text,
                                    reply_markup=student_search_result_by_fio_kb(students, action='open',
                                                                                 with_klass=True))
    else:
        text = 'Ученики не найдены.'
        await bot.edit_message_text(chat_id=msg.chat.id, message_id=msg_id, text=text, reply_markup=student_cancel_kb())


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "search_by_klass"))
async def callbacks_student_search_by_klass_step1_fab(callback: types.CallbackQuery,
                                                      callback_data: StudentCallbackFactory,
                                                      state: FSMContext):
    parallels = Parallel.select()
    text = '1️⃣Выберите параллель:'
    await callback.message.edit_text(text=text,
                                     reply_markup=student_parallel_choice_kb(parallels,
                                                                             action='choice_for_search_student'))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "choice_for_search_student"))
async def callbacks_student_search_by_klass_step2_fab(callback: types.CallbackQuery,
                                                      callback_data: StudentCallbackFactory,
                                                      state: FSMContext):
    parallel_id = callback_data.id
    parallel = Parallel.get(id=parallel_id)
    klasses = parallel.klasses
    text = '2️⃣Выберите класс'
    await callback.message.edit_text(text=text,
                                     reply_markup=student_klass_choice_kb(klasses, action='choice_for_search_student2'))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "choice_for_search_student2"))
async def student_search_by_klass_result_fab(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                             state: FSMContext):
    klass_id = callback_data.id
    page = callback_data.page
    klass = Klass(id=klass_id)
    students = klass.students

    text = f'Вот что удалось найти в базе учеников:'
    await callback.message.edit_text(text=text,
                                     reply_markup=student_search_result(students, action='open',
                                                                        parent_action='choice_for_search_student2',
                                                                        page=page, with_klass=False, klass_id=klass_id))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "open"))
async def student_card_fab(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                           state: FSMContext):
    await state.clear()
    student_id = callback_data.id
    student = Student.get(id=student_id)
    text = f'{student.fio} - {student.klass.name}'
    await callback.message.edit_text(text=text, reply_markup=current_student_kb(student_id))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "edit_name"))
async def callbacks_klass_edit_name(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                    state: FSMContext):
    student_id = callback_data.id
    await state.set_state(Content_add.student_edit_fio)
    await state.update_data(msg_id=callback.message.message_id)
    await state.update_data(student_id=student_id)
    text = callback.message.text + '\n🆕Введите новое имя ученика:'
    await callback.message.edit_text(text=text, reply_markup=back_to_student(student_id))


@student_router.message(Content_add.student_edit_fio)
async def klass_edit_name(msg: types.Message, state: FSMContext, bot: Bot):
    new_student_fio = msg.text.title()
    contex_data = await state.get_data()
    msg_id = contex_data.get('msg_id')
    student_id = contex_data.get('student_id')
    student = Student.get(id=student_id)
    student.fio = new_student_fio
    student.save()
    await msg.delete()
    await state.clear()
    text = f'{student.fio} - {student.klass.name}'
    await bot.edit_message_text(chat_id=msg.chat.id, message_id=msg_id, text=text,
                                reply_markup=current_student_kb(student_id))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "edit_klass"))
async def callbacks_student_change_klass_step1_fab(callback: types.CallbackQuery,
                                                   callback_data: StudentCallbackFactory,
                                                   state: FSMContext):
    student_id = callback_data.id
    await state.set_state(Content_add.student_edit_klass)
    await state.update_data(student_id=student_id)
    parallels = Parallel.select()
    text = '1️⃣Выберите параллель:'
    await callback.message.edit_text(text=text,
                                     reply_markup=student_parallel_choice_kb(parallels,
                                                                             action='choice_for_change'))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "choice_for_change"))
async def callbacks_student_change_klass_step2_fab(callback: types.CallbackQuery,
                                                   callback_data: StudentCallbackFactory,
                                                   state: FSMContext):
    parallel_id = callback_data.id
    parallel = Parallel.get(id=parallel_id)
    klasses = parallel.klasses
    text = '2️⃣Выберите класс'
    await callback.message.edit_text(text=text,
                                     reply_markup=student_klass_choice_kb(klasses, action='choice_for_change2'))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "choice_for_change2"))
async def callbacks_student_change_klass_ste3_fab(callback: types.CallbackQuery,
                                                  callback_data: StudentCallbackFactory,
                                                  state: FSMContext):
    klass_id = callback_data.id
    klass = Klass.get(id=klass_id)
    contex_data = await state.get_data()
    student_id = contex_data.get('student_id')
    student = Student.get(id=student_id)
    student.klass = klass
    student.save()
    text = f'{student.fio} - {student.klass.name}'
    await callback.message.edit_text(text=text, reply_markup=current_student_kb(student_id))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "delete"))
async def callbacks_student_delete(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                   state: FSMContext):
    student_id = callback_data.id
    text = callback.message.text + '\n❗️Уверены, что хотите удалить ученика?'
    await callback.message.edit_text(text=text, reply_markup=student_delete_confirm(student_id))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "delete_confirmed"))
async def callbacks_student_delete_confirmed(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                             state: FSMContext):
    student_id = callback_data.id
    student = Student.get(id=student_id)
    student_fio = student.fio
    student.delete_instance()
    text = f'Ученик {student_fio} удалён.'
    await callback.message.edit_text(text=text, reply_markup=main_students_kb())


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "book_issuance"))
async def callbacks_book_issuance(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                  state: FSMContext):
    student_id = callback_data.id
    await state.set_state(Content_add.await_books_qr_codes)
    await state.update_data(msg_id=callback.message.message_id)
    await state.update_data(student_id=student_id)
    await state.update_data({'codes': set()})
    text = 'Пришлите фото QR-кодов книг, которые нужно выдать и нажмите "Выдать"'
    await callback.message.edit_text(text=text, reply_markup=back_to_student(student_id))


@student_router.message(Content_add.await_books_qr_codes)
async def get_books_qr_codes(msg: types.Message, state: FSMContext, bot: Bot):
    photo_id = msg.photo[-1].file_id
    # photo = msg.photo
    photo = await bot.get_file(photo_id)
    img = qr_dir
    # await bot.get_file((img, f"{photo_id}.png"))
    file_path = os.path.join(img, f"{photo_id}.png")
    await bot.download(photo, destination=file_path)
    code = DecodeQR(file_path)
    contex_data = await state.get_data()
    codes = contex_data.get('codes') | {code}
    await state.update_data({'codes': codes})
    msg_id = contex_data.get('msg_id')
    student_id = contex_data.get('student_id')
    if code != '':
        text = f'✅ Отлично, я получил qr-код книги'
    else:
        text = '❌ Мне не удалось распознать QR-код'
    os.unlink(file_path)
    await msg.answer(text=text, reply_markup=student_book_issuance(student_id))


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "issuance_complete"))
async def callbacks_book_issuance_complete(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                           state: FSMContext):
    student_id = callback_data.id
    student = Student.get(id=student_id)
    # s1 = Staff.get(id=1)
    # print([s1.tg_id, callback.message.chat.id], callback.message.chat.id == s1.tg_id)
    staff = Staff.get(tg_id=str(callback.message.chat.id))
    contex_data = await state.get_data()
    codes = [code for code in contex_data.get('codes') if code != '']
    for code in codes:
        book = Book.get(id=code)
        book.free = False
        book.save()
        move = Book_movement.create(book=book, student=student, staff=staff,
                                    return_date=datetime.datetime.now() + datetime.timedelta(days=14))
    text = f'Книги выданы ученику {student.fio}'
    await callback.message.edit_text(text=text, reply_markup=student_cancel_kb())


@student_router.callback_query(StudentCallbackFactory.filter(F.action == "books"))
async def callbacks_students_books(callback: types.CallbackQuery, callback_data: StudentCallbackFactory,
                                           state: FSMContext):
    student_id = callback_data.id
    student = Student.get(id=student_id)
    text = f'Вот какие книги уже выданы ученику {student.fio}:'
    books = student.books
    for book in [book for book in books if book.is_return == False]:
        text += f'\n{book.book.book.name}'
    await callback.message.edit_text(text=text, reply_markup=back_to_student(student_id))
